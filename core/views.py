from datetime import date, timedelta, datetime
import calendar
from collections import defaultdict
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font, Alignment

from django import forms
from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth import authenticate, login, logout
from django.contrib.auth.decorators import login_required
from django.contrib.auth.forms import AuthenticationForm
from django.db.models import Sum
from django.urls import reverse_lazy
from django.views.generic import ListView, CreateView, UpdateView, DeleteView
from django.http import HttpResponse
from django.contrib import messages
from django.db import transaction


from .models import (
    Meals, MealRecipient, MonthlySummary,
    Category, StockItem, WeeklyStockMovement,
    FormG, FormH, RationAllowance
)
from .forms import (
    MealsForm, 
    
    CategoryForm, StockItemForm, WeeklyStockMovementForm,
    FormGForm
)


# Authentication Views
def login_view(request):
    """Handles user login using Django's built in AuthenticationForm"""
    if request.method == 'POST':
        form = AuthenticationForm(request, data=request.POST)
        if form.is_valid():
            user = form.get_user()
            login(request, user)
            return redirect('dashboard')
    else:
        form = AuthenticationForm()
    return render(request, 'registration/login.html', {'form': form})


def logout_view(request):
    """Logs out the user and redirects to login page"""
    logout(request)
    return redirect('login')


@login_required
def dashboard(request):
    """Dashboard view"""
    today = date.today()
    year = today.year
    week = today.isocalendar()[1]
    month_name = today.strftime('%B')
    
    # Quick stats for dashboard
    total_recipients = MealRecipient.objects.count()
    today_meals = Meals.objects.filter(mealDate=today).aggregate(
        total=Sum('quantity')
    )['total'] or 0
    
    recent_stock_movements = WeeklyStockMovement.objects.select_related(
        'stock_item', 'stock_item__category'
    ).order_by('-week_number')[:5]

    return render(request, "core/main.html", {
        "year": year,
        "month": today.month,
        "week": week,
        "month_name": month_name,
        "total_recipients": total_recipients,
        "today_meals": today_meals,
        "recent_stock_movements": recent_stock_movements,
    })


def home(request):
    """Home page view"""
    return render(request, 'core/home.html')



# Meal Views
# Meal Views
class MealListView(ListView):
    model = Meals
    template_name = "core/meals_list.html"
    context_object_name = "meals"
    ordering = ["-mealDate"]
    paginate_by = 20

class MealCreateView(CreateView):
    model = Meals
    form_class = MealsForm
    template_name = "core/meal_form.html"
    success_url = reverse_lazy("meals_list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form_title"] = "Add New Meal"
        return context

    def form_valid(self, form):
        messages.success(self.request, "Meal added successfully!")
        return super().form_valid(form)

class MealUpdateView(UpdateView):
    model = Meals
    form_class = MealsForm
    template_name = "core/meal_form.html"
    success_url = reverse_lazy("meals_list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form_title"] = "Edit Meal"
        return context

    def form_valid(self, form):
        messages.success(self.request, "Meal updated successfully!")
        return super().form_valid(form)

class MealDeleteView(DeleteView):
    model = Meals
    template_name = "core/meal_confirm_delete.html"
    success_url = reverse_lazy("meals_list")
    context_object_name = "meal"

    def delete(self, request, *args, **kwargs):
        messages.success(request, "Meal deleted successfully!")
        return super().delete(request, *args, **kwargs)

# Simple weekly summary view (read-only)
@login_required
def weekly_meals_overview(request):
    """Simple overview of weekly meals - read only"""
    today = date.today()
    year = request.GET.get('year', today.year)
    week = request.GET.get('week', today.isocalendar()[1])
    
    try:
        year = int(year)
        week = int(week)
        start_date = date.fromisocalendar(year, week, 1)  # Monday
        end_date = start_date + timedelta(days=6)  # Sunday
    except (ValueError, TypeError):
        start_date = date.fromisocalendar(today.year, today.isocalendar()[1], 1)
        end_date = start_date + timedelta(days=6)
    
    # Get all meals for the week
    meals = Meals.objects.filter(
        mealDate__range=[start_date, end_date]
    ).select_related('mealsFor').order_by('mealDate', 'mealsFor__name')
    
    # Group by recipient and date
    meals_by_recipient = {}
    for meal in meals:
        recipient_name = meal.mealsFor.name
        if recipient_name not in meals_by_recipient:
            meals_by_recipient[recipient_name] = {}
        
        date_str = meal.mealDate.strftime('%Y-%m-%d')
        if date_str not in meals_by_recipient[recipient_name]:
            meals_by_recipient[recipient_name][date_str] = {}
        
        meals_by_recipient[recipient_name][date_str][meal.mealCategory] = meal.quantity
    
    # Create dates for the template
    dates = []
    for i in range(7):
        current_date = start_date + timedelta(days=i)
        dates.append({
            'date': current_date,
            'day_name': current_date.strftime('%A'),
            'date_short': current_date.strftime('%m/%d')
        })
    
    context = {
        'meals_by_recipient': meals_by_recipient,
        'dates': dates,
        'start_date': start_date,
        'end_date': end_date,
        'year': year,
        'week': week,
        'meal_types': dict(Meals.CATEGORY_CHOICES),
    }
    return render(request, 'core/weekly_meals_overview.html', context)


# Meal Recipient Views
class RecipientListView(ListView):
    model = MealRecipient
    template_name = "core/recipients_list.html"
    context_object_name = "recipients"
    ordering = ["name"]


class RecipientCreateView(CreateView):
    model = MealRecipient
    fields = ["name"]
    template_name = "core/recipient_form.html"
    success_url = reverse_lazy("recipients_list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form_title"] = "Add New Recipient"
        return context

    def form_valid(self, form):
        messages.success(self.request, "Recipient added successfully!")
        return super().form_valid(form)


class RecipientUpdateView(UpdateView):
    model = MealRecipient
    fields = ["name"]
    template_name = "core/recipient_form.html"
    success_url = reverse_lazy("recipients_list")

    def get_context_data(self, **kwargs):
        context = super().get_context_data(**kwargs)
        context["form_title"] = "Edit Recipient"
        return context

    def form_valid(self, form):
        messages.success(self.request, "Recipient updated successfully!")
        return super().form_valid(form)


class RecipientDeleteView(DeleteView):
    model = MealRecipient
    template_name = "core/meal_confirm_delete.html"
    success_url = reverse_lazy("recipients_list")
    context_object_name = "recipient"

    def delete(self, request, *args, **kwargs):
        messages.success(request, "Recipient deleted successfully!")
        return super().delete(request, *args, **kwargs)

@login_required
def weekly_summary(request, year, week):
    """Generates a weekly summary table with days as rows and meal types as columns"""
    start_date = date.fromisocalendar(year, week, 1)
    end_date = start_date + timedelta(days=6)

    recipients = MealRecipient.objects.all()
    meal_types = dict(Meals.CATEGORY_CHOICES)

    # Create days list with date objects for template
    days = []
    for day in range(7):
        current_date = start_date + timedelta(days=day)
        days.append({
            'date': current_date,
            'day_name': current_date.strftime('%A'),
            'date_short': current_date.strftime('%m/%d')
        })

    # Reorganize data for the new structure
    summary = {}
    daily_totals = [0] * 7
    recipient_totals = {}
    grand_total = 0

    for recipient in recipients:
        recipient_totals[recipient.name] = {}
        for code in meal_types.keys():
            recipient_totals[recipient.name][code] = 0

        summary[recipient.name] = []
        for day in range(7):
            meal_date = start_date + timedelta(days=day)
            day_data = {}
            day_total = 0
            for code, _ in Meals.CATEGORY_CHOICES:
                qty = Meals.objects.filter(
                    mealDate=meal_date,
                    mealsFor=recipient,
                    mealCategory=code
                ).aggregate(Sum("quantity"))["quantity__sum"] or 0
                day_data[code] = qty
                day_total += qty
                recipient_totals[recipient.name][code] += qty
            summary[recipient.name].append(day_data)
            daily_totals[day] += day_total
            grand_total += day_total

    context = {
        "summary": summary,
        "recipients": recipients,
        "daily_totals": daily_totals,
        "recipient_totals": recipient_totals,
        "grand_total": grand_total,
        "start_date": start_date,
        "end_date": end_date,
        "meal_types": meal_types,
        "days": days,
        "year": year,
        "week": week,
    }
    return render(request, "core/weekly_summary.html", context)

# Monthly Summary Views
@login_required
def monthly_summary(request, year=None, month=None):
    """Generates monthly meal summary for all recipients"""
    today = date.today()
    year = year or today.year
    month = month or today.month

    try:
        num_days = calendar.monthrange(year, month)[1]
        first_day = date(year, month, 1)
        last_day = date(year, month, num_days)
    except (ValueError, TypeError):
        messages.error(request, "Invalid month or year provided.")
        return redirect('monthly_summary_select')

    recipients = MealRecipient.objects.all()
    meal_types = dict(Meals.CATEGORY_CHOICES)

    summary = []
    for recipient in recipients:
        row = {"recipient": recipient.name, "meals": []}
        total = 0
        for day in range(1, num_days + 1):
            meal_date = date(year, month, day)
            day_data = {}
            for code, _ in Meals.CATEGORY_CHOICES:
                qty = Meals.objects.filter(
                    mealDate=meal_date,
                    mealsFor=recipient,
                    mealCategory=code
                ).aggregate(Sum("quantity"))["quantity__sum"] or 0
                day_data[code] = qty
                total += qty
            row["meals"].append(day_data)
        row["total"] = total
        summary.append(row)

    return render(request, "core/monthly_summary.html", {
        "summary": summary,
        "year": year,
        "month": month,
        "meal_types": meal_types,
        "num_days": num_days,
        "month_name": first_day.strftime('%B'),
    })


class MonthSelectForm(forms.Form):
    """Form to select a month and year for monthly summary"""
    year = forms.IntegerField(
        initial=date.today().year,
        min_value=2000,
        max_value=2100,
        widget=forms.NumberInput(attrs={"class": "form-control"})
    )
    month = forms.IntegerField(
        min_value=1, max_value=12,
        initial=date.today().month,
        widget=forms.NumberInput(attrs={"class": "form-control"})
    )


@login_required
def monthly_summary_select(request):
    """View to select month/year and redirect to monthly summary"""
    if request.method == "POST":
        form = MonthSelectForm(request.POST)
        if form.is_valid():
            year = form.cleaned_data["year"]
            month = form.cleaned_data["month"]
            return redirect("monthly_summary", year=year, month=month)
    else:
        form = MonthSelectForm()

    return render(request, "core/monthly_summary_select.html", {"form": form})


# Stock Management Views
@login_required
def category_list(request):
    """Category list view"""
    categories = Category.objects.all().order_by("category_no")
    return render(request, "core/categories.html", {"categories": categories})


@login_required
def category_add(request):
    """Add category view"""
    if request.method == "POST":
        form = CategoryForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Category added successfully!")
            return redirect("categories")
    else:
        form = CategoryForm()
    return render(request, "core/form.html", {"form": form, "title": "Add Category"})


@login_required
def category_edit(request, pk):
    """Edit category view"""
    category = get_object_or_404(Category, pk=pk)
    if request.method == "POST":
        form = CategoryForm(request.POST, instance=category)
        if form.is_valid():
            form.save()
            messages.success(request, "Category updated successfully!")
            return redirect("categories")
    else:
        form = CategoryForm(instance=category)
    return render(request, "core/form.html", {"form": form, "title": "Edit Category"})


@login_required
def category_delete(request, pk):
    """Delete category view"""
    category = get_object_or_404(Category, pk=pk)
    if request.method == "POST":
        category.delete()
        messages.success(request, "Category deleted successfully!")
        return redirect("categories")
    return render(request, "core/confirm_delete.html", {"object": category, "title": "Delete Category"})


@login_required
def stock_list(request):
    """Stock item list view"""
    stock_items = StockItem.objects.select_related("category").all().order_by("category__category_no", "name")
    return render(request, "core/Stock_list.html", {"stock_items": stock_items})


@login_required
def stock_item_add(request):
    """Add stock item view"""
    if request.method == "POST":
        form = StockItemForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Stock item added successfully!")
            return redirect("stock_list")
    else:
        form = StockItemForm()
    return render(request, "core/form.html", {"form": form, "title": "Add Stock Item"})


@login_required
def stock_item_edit(request, pk):
    """Edit stock item view"""
    item = get_object_or_404(StockItem, pk=pk)
    if request.method == "POST":
        form = StockItemForm(request.POST, instance=item)
        if form.is_valid():
            form.save()
            messages.success(request, "Stock item updated successfully!")
            return redirect("stock_list")
    else:
        form = StockItemForm(instance=item)
    return render(request, "core/form.html", {"form": form, "title": "Edit Stock Item"})


@login_required
def stock_item_delete(request, pk):
    """Delete stock item view"""
    item = get_object_or_404(StockItem, pk=pk)
    if request.method == "POST":
        item.delete()
        messages.success(request, "Stock item deleted successfully!")
        return redirect("stock_list")
    return render(request, "core/confirm_delete.html", {"object": item, "title": "Delete Stock Item"})


@login_required
def stock_movement_list(request):
    """Stock movement list view"""
    movements = WeeklyStockMovement.objects.select_related("stock_item", "stock_item__category").all()
    return render(request, "core/stock_movement_list.html", {"movements": movements})


@login_required
def stock_movement_add(request):
    """Add stock movement view"""
    if request.method == "POST":
        form = WeeklyStockMovementForm(request.POST)
        if form.is_valid():
            movement = form.save(commit=False)
            
            # Auto-calculate cost if not provided
            if not movement.cost and hasattr(movement, 'unit_price') and movement.unit_price and movement.total_received:
                movement.cost = movement.unit_price * movement.total_received
            
            movement.save()
            
            # Set next week's start stock to this week's end stock
            if hasattr(movement, 'end_stock'):
                next_week = movement.week_number + 1
                if next_week <= 53:
                    next_movement, created = WeeklyStockMovement.objects.get_or_create(
                        stock_item=movement.stock_item,
                        week_number=next_week,
                        defaults={'start_stock': movement.end_stock}
                    )
                    if not created and hasattr(next_movement, 'start_stock'):
                        next_movement.start_stock = movement.end_stock
                        next_movement.save()
            
            messages.success(request, "Stock movement added successfully!")
            return redirect("stock_movement_list")
    else:
        # Auto-set start stock from previous week
        week_number = request.GET.get('week')
        stock_item_id = request.GET.get('stock_item')
        initial_data = {}
        
        if week_number and stock_item_id:
            try:
                prev_week = int(week_number) - 1
                if prev_week > 0:
                    prev_movement = WeeklyStockMovement.objects.filter(
                        week_number=prev_week,
                        stock_item_id=stock_item_id
                    ).first()
                    if prev_movement and hasattr(prev_movement, 'end_stock'):
                        initial_data['start_stock'] = prev_movement.end_stock
            except ValueError:
                pass
                
        form = WeeklyStockMovementForm(initial=initial_data)
    
    return render(request, "core/form.html", {"form": form, "title": "Add Stock Movement"})


@login_required
def stock_movement_edit(request, pk):
    """Edit stock movement view"""
    movement = get_object_or_404(WeeklyStockMovement, pk=pk)
    if request.method == "POST":
        form = WeeklyStockMovementForm(request.POST, instance=movement)
        if form.is_valid():
            form.save()
            messages.success(request, "Stock movement updated successfully!")
            return redirect("stock_movement_list")
    else:
        form = WeeklyStockMovementForm(instance=movement)
    return render(request, "core/form.html", {"form": form, "title": "Edit Stock Movement"})


@login_required
def stock_movement_delete(request, pk):
    """Delete stock movement view"""
    movement = get_object_or_404(WeeklyStockMovement, pk=pk)
    if request.method == "POST":
        movement.delete()
        messages.success(request, "Stock movement deleted successfully!")
        return redirect("stock_movement_list")
    return render(request, "core/confirm_delete.html", {"object": movement, "title": "Delete Stock Movement"})


@login_required
def stock_weekly_summary(request):
    """Generates weekly stock summary based on selected week"""
    weeks = WeeklyStockMovement.objects.values_list("week_number", flat=True).distinct().order_by("week_number")
    selected_week = int(request.GET.get("week_number", weeks.first() if weeks else 0))

    summary = []
    totals = {}

    if selected_week:
        summary = (
            WeeklyStockMovement.objects
            .filter(week_number=selected_week)
            .values("stock_item__category__description")
            .annotate(
                total_received=Sum("total_received"),
                total_issued=Sum("total_issued"),
                total_extern=Sum("extern_issues"),
                total_cost=Sum("cost")
            )
            .order_by("stock_item__category__description")
        )

        totals = WeeklyStockMovement.objects.filter(week_number=selected_week).aggregate(
            total_received=Sum("total_received"),
            total_issued=Sum("total_issued"),
            total_extern=Sum("extern_issues"),
            total_cost=Sum("cost")
        )

    return render(request, "core/Stock Weekly_summary.html", {
        "summary": summary,
        "totals": totals,
        "weeks": weeks,
        "selected_week": selected_week,
    })


# Export Views
@login_required
def weekly_summary_preview(request, year, week):
    """Weekly summary preview for export"""
    start_date = date.fromisocalendar(year, week, 1)
    end_date = start_date + timedelta(days=6)

    meals = Meals.objects.filter(mealDate__range=[start_date, end_date]).select_related('mealsFor')

    context = {
        "year": year,
        "week": week,
        "meals": meals,
        "start_date": start_date,
        "end_date": end_date,
        "date_range": f"{start_date.strftime('%b %d, %Y')} - {end_date.strftime('%b %d, %Y')}",
    }
    return render(request, "core/weekly_summary_preview.html", context)


@login_required
def monthly_summary_preview(request, year, month):
    """Monthly summary preview for export"""
    first_day = date(year, month, 1)
    last_day = date(year, month + 1, 1) - timedelta(days=1) if month < 12 else date(year, 12, 31)

    meals = Meals.objects.filter(mealDate__range=[first_day, last_day])

    context = {
        "year": year,
        "month": month,
        "meals": meals,
        "month_name": first_day.strftime('%B'),
        "date_range": f"{first_day.strftime('%d %b')} - {last_day.strftime('%d %b %Y')}",
    }
    return render(request, "core/monthly_summary_preview.html", context)


@login_required
def export_weekly_summary_excel(request, year, week):
    """Export weekly summary to Excel"""
    start_date = date.fromisocalendar(year, week, 1)
    end_date = start_date + timedelta(days=6)

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Week_{week}_{year}"

    # Create headers with actual dates
    headers = ["Recipient"]
    meal_types = [code for code, _ in Meals.CATEGORY_CHOICES]
    for day in range(7):
        current_date = start_date + timedelta(days=day)
        for mt in meal_types:
            headers.append(f"{current_date.strftime('%d %b')} {mt}")
    headers.append("Total")
    ws.append(headers)

    recipients = MealRecipient.objects.all()
    for recipient in recipients:
        row = [recipient.name]
        total = 0
        for day in range(7):
            meal_date = start_date + timedelta(days=day)
            for code, _ in Meals.CATEGORY_CHOICES:
                qty = Meals.objects.filter(
                    mealDate=meal_date,
                    mealsFor=recipient,
                    mealCategory=code
                ).aggregate(total=Sum("quantity"))["total"] or 0
                row.append(qty)
                total += qty
        row.append(total)
        ws.append(row)

    # Adjust column widths
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename=Weekly_Summary_Week{week}_{year}.xlsx'
    wb.save(response)
    return response


@login_required
def export_monthly_summary_excel(request, year=None, month=None):
    """Export monthly summary to Excel"""
    year = year or date.today().year
    month = month or date.today().month

    first_day = date(year, month, 1)
    last_day = date(year, month + 1, 1) - timedelta(days=1) if month < 12 else date(year, 12, 31)
    num_days = (last_day - first_day).days + 1

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{first_day.strftime('%B')}-{year} Summary"

    # Headers
    headers = ["Recipient"]
    recipients = MealRecipient.objects.all()
    meal_types = [code for code, _ in Meals.CATEGORY_CHOICES]

    for day in range(num_days):
        current_date = first_day + timedelta(days=day)
        for mt in meal_types:
            headers.append(f"{current_date.strftime('%d %b')} {mt}")
    headers.append("Total")
    ws.append(headers)

    # Fill data
    for recipient in recipients:
        row = [recipient.name]
        total = 0
        for day in range(num_days):
            meal_date = first_day + timedelta(days=day)
            for code, _ in Meals.CATEGORY_CHOICES:
                qty = Meals.objects.filter(
                    mealDate=meal_date,
                    mealsFor=recipient,
                    mealCategory=code
                ).aggregate(total=Sum("quantity"))["total"] or 0
                row.append(qty)
                total += qty
        row.append(total)
        ws.append(row)

    # Adjust column widths
    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    # Response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename=Monthly_Summary_{first_day.strftime("%b")}_{year}.xlsx'
    wb.save(response)
    return response


@login_required
def stock_movement_preview(request):
    """Stock movement preview for export"""
    movements = WeeklyStockMovement.objects.select_related(
        "stock_item", "stock_item__category"
    ).all()

    categories = Category.objects.all()
    category_totals = []
    for cat in categories:
        totals = WeeklyStockMovement.objects.filter(
            stock_item__category=cat
        ).aggregate(
            total_received=Sum("total_received"),
            total_issued=Sum("total_issued"),
            total_extern=Sum("extern_issues"),
            total_cost=Sum("cost")
        )
        totals["description"] = cat.description
        category_totals.append(totals)

    grand_totals = WeeklyStockMovement.objects.aggregate(
        total_received=Sum("total_received"),
        total_issued=Sum("total_issued"),
        total_extern=Sum("extern_issues"),
        total_cost=Sum("cost")
    )

    return render(request, "core/stock_movement_preview.html", {
        "movements": movements,
        "category_totals": category_totals,
        "grand_totals": grand_totals
    })


@login_required
def export_stock_excel(request):
    """Export stock data to Excel"""
    wb = openpyxl.Workbook()

    # Sheet 1: Stock Movement
    ws1 = wb.active
    ws1.title = "Stock Movement"

    headers = [
        "Week Number", "Stock Item", "Category",
        "Total Received", "Total Issued", "External Issues", "Cost"
    ]
    ws1.append(headers)

    movements = WeeklyStockMovement.objects.select_related("stock_item", "stock_item__category").all()
    
    for m in movements:
        ws1.append([
            m.week_number,
            m.stock_item.name,
            m.stock_item.category.description,
            float(m.total_received),
            float(m.total_issued),
            float(m.extern_issues),
            float(m.cost),
        ])

    # Totals per Category
    categories = Category.objects.all()
    ws1.append([])
    ws1.append(["Totals per Category"])
    ws1.append(["Category", "Total Received", "Total Issued", "External Issues", "Total Cost"])

    for cat in categories:
        cat_totals = WeeklyStockMovement.objects.filter(stock_item__category=cat).aggregate(
            total_received=Sum("total_received"),
            total_issued=Sum("total_issued"),
            total_extern=Sum("extern_issues"),
            total_cost=Sum("cost")
        )
        ws1.append([
            cat.description,
            float(cat_totals["total_received"] or 0),
            float(cat_totals["total_issued"] or 0),
            float(cat_totals["total_extern"] or 0),
            float(cat_totals["total_cost"] or 0),
        ])

    # Grand Totals
    grand_totals = WeeklyStockMovement.objects.aggregate(
        total_received=Sum("total_received"),
        total_issued=Sum("total_issued"),
        total_extern=Sum("extern_issues"),
        total_cost=Sum("cost")
    )
    ws1.append([])
    ws1.append([
        "GRAND TOTAL",
        "",
        "",
        float(grand_totals["total_received"] or 0),
        float(grand_totals["total_issued"] or 0),
        float(grand_totals["total_extern"] or 0),
        float(grand_totals["total_cost"] or 0),
    ])

    # Adjust column widths
    for col in ws1.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        ws1.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    # Sheet 2: Stock Items
    ws2 = wb.create_sheet(title="Stock Items")
    items = StockItem.objects.select_related("category").all()
    ws2.append(["Name", "Unit", "Size", "Category"])
    for item in items:
        ws2.append([item.name, item.unit, item.size or "", item.category.description])
    for col in ws2.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        ws2.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    # Sheet 3: Categories
    ws3 = wb.create_sheet(title="Categories")
    categories = Category.objects.all().order_by("category_no")
    ws3.append(["Category No", "Description"])
    for cat in categories:
        ws3.append([cat.category_no, cat.description])
    for col in ws3.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        ws3.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = 'attachment; filename="Stock_Summary.xlsx"'
    wb.save(response)
    return response


# Form G Views
@login_required
def form_g_list(request):
    """Displays all Form G summaries"""
    data = FormG.objects.select_related('category').all().order_by("-year", "-month", "category__description")
    return render(request, "core/form_g_list.html", {"data": data})


@login_required
def form_g_add(request):
    """Add a new Form G record"""
    if request.method == "POST":
        form = FormGForm(request.POST)
        if form.is_valid():
            form.save()
            messages.success(request, "Form G record added successfully!")
            return redirect("form_g_list")
    else:
        form = FormGForm()

    return render(request, "core/form_g_form.html", {"form": form, "title": "Add Form G Record"})


@login_required
def form_g_monthly_preview(request, year, month):
    """Form G monthly preview"""
    entries = FormG.objects.filter(year=year, month=month).order_by("category__description")

    context = {
        "year": year,
        "month": month,
        "entries": entries,
        "month_name": date(year, month, 1).strftime("%B"),
    }
    return render(request, "core/form_g_monthly_preview.html", context)


@login_required
def export_form_g_monthly_excel(request, year=None, month=None):
    """Export Form G to Excel"""
    year = year or date.today().year
    month = month or date.today().month

    entries = FormG.objects.filter(year=year, month=month).order_by("category__description")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{date(year, month, 1).strftime('%B')}-{year} Form G"

    headers = [
        "Category", "Annual Budget", "Monthly Budget",
        "Week 1 Expense", "Week 2 Expense", "Week 3 Expense",
        "Week 4 Expense", "Week 5 Expense", "Expense for Month", "Underspent"
    ]
    ws.append(headers)

    for entry in entries:
        row = [
            entry.category.description,
            float(entry.annual_budget),
            float(entry.monthly_budget),
            float(entry.week1_expense),
            float(entry.week2_expense),
            float(entry.week3_expense),
            float(entry.week4_expense),
            float(entry.week5_expense),
            float(entry.expense_for_month),
            float(entry.underspent),
        ]
        ws.append(row)

    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename=FormG_Summary_{month}_{year}.xlsx'
    wb.save(response)
    return response


# Form H Views
@login_required
def generate_formh_summary(request, year, month):
    """Generate Form H summary"""
    categories = Category.objects.all()
    formh_entries = []

    for category in categories:
        try:
            g_entry = FormG.objects.get(category=category, year=year, month=month)
        except FormG.DoesNotExist:
            continue

        allowance = RationAllowance.objects.filter(category=category).first()
        
        # FIX: Use expense_for_month instead of total_expense_for_month
        total_usage = g_entry.expense_for_month / 100  # Convert to appropriate units if needed
        avg_per_person_per_day = round(total_usage / 30, 3) if total_usage else 0

        formh, _ = FormH.objects.update_or_create(
            category=category, year=year, month=month,
            defaults={
                "week1_usage": g_entry.week1_expense or 0,
                "week2_usage": g_entry.week2_expense or 0,
                "week3_usage": g_entry.week3_expense or 0,
                "week4_usage": g_entry.week4_expense or 0,
                "week5_usage": g_entry.week5_expense or 0,
                "total_usage": total_usage,
                "avg_per_person_per_day": avg_per_person_per_day,
            },
        )

        out_of_range = False
        if allowance and avg_per_person_per_day:
            out_of_range = (
                avg_per_person_per_day < allowance.min_allowance
                or avg_per_person_per_day > allowance.max_allowance
            )

        formh_entries.append({
            "formh": formh,
            "allowance": allowance,
            "out_of_range": out_of_range,
        })

    return render(request, "core/formh_summary.html", {
        "entries": formh_entries,
        "month": month,
        "year": year,
    })

@login_required
def form_h_monthly_preview(request, year, month):
    """Form H monthly preview"""
    entries = FormH.objects.filter(year=year, month=month).order_by("category__description")

    context = {
        "year": year,
        "month": month,
        "entries": entries,
        "month_name": date(year, month, 1).strftime("%B"),
    }
    return render(request, "core/form_h_monthly_preview.html", context)


@login_required
def export_form_h_monthly_excel(request, year=None, month=None):
    """Export Form H to Excel"""
    year = year or date.today().year
    month = month or date.today().month

    entries = FormH.objects.filter(year=year, month=month).order_by("category__description")

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"{date(year, month, 1).strftime('%B')}-{year} Form H"

    headers = [
        "Category",
        "Week 1 Usage", "Week 2 Usage", "Week 3 Usage",
        "Week 4 Usage", "Week 5 Usage", "Total Usage",
        "Adjusted Total Usage", "Avg per Person/Day",
        "Allowance Min", "Allowance Max", "Unit/Per", "Status"
    ]
    ws.append(headers)

    for entry in entries:
        allowance = entry.category.rationallowance_set.first()  # may be None
        row = [
            entry.category.description,
            float(entry.week1_usage),
            float(entry.week2_usage),
            float(entry.week3_usage),
            float(entry.week4_usage),
            float(entry.week5_usage),
            float(entry.total_usage),
            float(entry.adjusted_total_usage),
            float(entry.avg_per_person_per_day),
            float(allowance.min_allowance) if allowance else None,
            float(allowance.max_allowance) if allowance else None,
            f"{allowance.unit}/{allowance.per}" if allowance else None,
            "Out of Range" if allowance and (entry.avg_per_person_per_day < allowance.min_allowance or entry.avg_per_person_per_day > allowance.max_allowance) else "OK",
        ]
        ws.append(row)

    for col in ws.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        ws.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename=FormH_Summary_{month}_{year}.xlsx'
    wb.save(response)
    return response


@login_required
def annual_summary(request, year=None):
    """FORM X - Annual feeding, expenditure and ration summary"""
    year = year or date.today().year
    
    # Calculate annual data directly from Meals model
    meals_queryset = Meals.objects.filter(mealDate__year=year)
    total_meals = meals_queryset.aggregate(total=Sum('quantity'))['total'] or 0
    
    # Calculate expenses manually from FormG
    form_g_data = FormG.objects.filter(year=year)
    total_expenses = 0
    total_budget = 0
    
    for form_g in form_g_data:
        # Calculate expense_for_month manually
        expense = sum([
            form_g.week1_expense or 0,
            form_g.week2_expense or 0, 
            form_g.week3_expense or 0,
            form_g.week4_expense or 0,
            form_g.week5_expense or 0
        ])
        total_expenses += expense
        total_budget += form_g.monthly_budget or 0
    
    avg_cost_per_meal = 0
    if total_meals > 0:
        avg_cost_per_meal = total_expenses / total_meals
    
    # Get monthly breakdown from actual Meals data
    monthly_breakdown = []
    for month in range(1, 13):
        # Get meals for this month
        month_meals = meals_queryset.filter(mealDate__month=month).aggregate(
            total=Sum('quantity')
        )['total'] or 0
        
        # Calculate monthly expenses from FormG
        month_expenses = 0
        month_form_g = FormG.objects.filter(year=year, month=month)
        for form_g in month_form_g:
            month_expenses += sum([
                form_g.week1_expense or 0,
                form_g.week2_expense or 0,
                form_g.week3_expense or 0, 
                form_g.week4_expense or 0,
                form_g.week5_expense or 0
            ])
        
        monthly_breakdown.append({
            'month': month,
            'month_name': date(year, month, 1).strftime('%B'),
            'meals': month_meals,
            'expenses': month_expenses,
        })
    
    context = {
        'year': year,
        'total_meals': total_meals,
        'total_expenses': total_expenses,
        'total_budget': total_budget,
        'avg_cost_per_meal': avg_cost_per_meal,
        'monthly_breakdown': monthly_breakdown,
    }
    
    return render(request, 'core/annual_summary.html', context)
   

@login_required
def stock_variance_report(request):
    """Report showing stock variances that need attention"""
    movements = WeeklyStockMovement.objects.select_related('stock_item', 'stock_item__category').order_by('-week_number', 'stock_item__name')
    
    # Calculate variance manually since we don't have the new fields in database yet
    movements_with_variance = []
    for movement in movements:
        # Calculate variance manually
        calculated_end_stock = (movement.total_received or 0) - (movement.total_issued or 0) - (movement.extern_issues or 0)
       
        variance = 0  
        
   
        needs_attention = (
            (movement.total_received > 0 and movement.total_issued == 0) or
            (movement.total_issued > movement.total_received) or
            movement.extern_issues > movement.total_received
        )
        
        movements_with_variance.append({
            'movement': movement,
            'variance': variance,
            'needs_attention': needs_attention,
            'calculated_end_stock': calculated_end_stock
        })
    
    context = {
        'movements_with_variance': movements_with_variance,
        'title': 'Stock Variance Report'
    }
    return render(request, 'core/stock_variance_report.html', context)

@login_required
def export_annual_summary_excel(request, year=None):
    """Export annual summary to Excel"""
    year = year or date.today().year
    
    # Calculate annual data directly from Meals model (same as annual_summary view)
    meals_queryset = Meals.objects.filter(mealDate__year=year)
    total_meals = meals_queryset.aggregate(total=Sum('quantity'))['total'] or 0
    
    # Calculate expenses manually from FormG
    form_g_data = FormG.objects.filter(year=year)
    total_expenses = 0
    total_budget = 0
    
    for form_g in form_g_data:
        # Calculate expense_for_month manually
        expense = sum([
            form_g.week1_expense or 0,
            form_g.week2_expense or 0, 
            form_g.week3_expense or 0,
            form_g.week4_expense or 0,
            form_g.week5_expense or 0
        ])
        total_expenses += expense
        total_budget += form_g.monthly_budget or 0
    
    avg_cost_per_meal = 0
    if total_meals > 0:
        avg_cost_per_meal = total_expenses / total_meals
    
    # Get monthly breakdown from actual Meals data
    monthly_breakdown = []
    for month in range(1, 13):
        # Get meals for this month
        month_meals = meals_queryset.filter(mealDate__month=month).aggregate(
            total=Sum('quantity')
        )['total'] or 0
        
        # Calculate monthly expenses from FormG
        month_expenses = 0
        month_form_g = FormG.objects.filter(year=year, month=month)
        for form_g in month_form_g:
            month_expenses += sum([
                form_g.week1_expense or 0,
                form_g.week2_expense or 0,
                form_g.week3_expense or 0, 
                form_g.week4_expense or 0,
                form_g.week5_expense or 0
            ])
        
        monthly_breakdown.append({
            'month': month,
            'month_name': date(year, month, 1).strftime('%B'),
            'meals': month_meals,
            'expenses': month_expenses,
        })

    # Create Excel workbook
    wb = openpyxl.Workbook()
    
    # Sheet 1: Annual Summary
    ws1 = wb.active
    ws1.title = f"Annual Summary {year}"
    
    # Title and headers
    title_font = Font(size=16, bold=True)
    header_font = Font(bold=True)
    
    ws1.merge_cells('A1:D1')
    ws1['A1'] = f"ANNUAL FEEDING, EXPENDITURE AND RATION SUMMARY - {year}"
    ws1['A1'].font = title_font
    ws1['A1'].alignment = Alignment(horizontal='center')
    
    # Summary section
    ws1.append([])
    ws1.append(["ANNUAL SUMMARY"])
    ws1.append(["Total Meals Served:", total_meals])
    ws1.append(["Total Expenses:", f"R{total_expenses:,.2f}"])
    ws1.append(["Total Budget:", f"R{total_budget:,.2f}"])
    ws1.append(["Average Cost Per Meal:", f"R{avg_cost_per_meal:.2f}"])
    
    # Monthly breakdown section
    ws1.append([])
    ws1.append(["MONTHLY BREAKDOWN"])
    ws1.append(["Month", "Meals Served", "Expenses", "Cost Per Meal"])
    
    for month_data in monthly_breakdown:
        month_cost_per_meal = 0
        if month_data['meals'] > 0:
            month_cost_per_meal = month_data['expenses'] / month_data['meals']
        
        ws1.append([
            month_data['month_name'],
            month_data['meals'],
            f"R{month_data['expenses']:,.2f}",
            f"R{month_cost_per_meal:.2f}"
        ])
    
    # Format headers
    for row in ws1.iter_rows(min_row=1, max_row=ws1.max_row):
        for cell in row:
            if cell.value in ["ANNUAL SUMMARY", "MONTHLY BREAKDOWN"] or cell.row <= 6:
                cell.font = header_font
    
    # Adjust column widths
    for col in ws1.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        ws1.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    # Sheet 2: Detailed Monthly Data
    ws2 = wb.create_sheet(title=f"Monthly Details {year}")
    
    # Headers for detailed monthly data
    headers = ["Month", "Category", "Week 1", "Week 2", "Week 3", "Week 4", "Week 5", "Monthly Total", "Monthly Budget"]
    ws2.append(headers)
    
    # Add Form G data for each month
    for month in range(1, 13):
        form_g_entries = FormG.objects.filter(year=year, month=month).order_by('category__description')
        
        for entry in form_g_entries:
            ws2.append([
                date(year, month, 1).strftime('%B'),
                entry.category.description,
                float(entry.week1_expense or 0),
                float(entry.week2_expense or 0),
                float(entry.week3_expense or 0),
                float(entry.week4_expense or 0),
                float(entry.week5_expense or 0),
                float(entry.expense_for_month),
                float(entry.monthly_budget or 0)
            ])
    
    # Format headers for sheet 2
    for cell in ws2[1]:
        cell.font = header_font
    
    # Adjust column widths for sheet 2
    for col in ws2.columns:
        max_length = max(len(str(cell.value)) for cell in col if cell.value)
        ws2.column_dimensions[get_column_letter(col[0].column)].width = max_length + 2

    # Create response
    response = HttpResponse(
        content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    response['Content-Disposition'] = f'attachment; filename="Annual_Summary_{year}.xlsx"'
    wb.save(response)
    return response